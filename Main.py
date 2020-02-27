from tkinter import*
import tkinter.messagebox
from threading import Thread
import openpyxl
from webbrowser import open as openlink
import time 
import random
def _from_rgb(rgb):#changes rbg values to tkinter friendly colors
    return "#%02x%02x%02x" % rgb
def times():
	return time.strftime("%d/%m/%Y-%H:%M:%S")
class tk():
    def __init__(self,root):
        self.root=root
        root.title('Allied High School®')
        self.img01=PhotoImage(file='bg.png')
        self.labelsframe=Frame(self.root)
        self.h=root.winfo_screenheight()
        self.labelsframe.place(x=0,y=100)
        self.labelsframe=Frame(self.root,)
        self.labelsframe.place(x=0,y=100)
        self.root.geometry('1300x'+str(self.h-10))
        self.labelbg=Label(self.root,image=self.img01)
        self.labelbg.place(x=0 ,y=0, width=1300, height=995)
        self.img00=PhotoImage(file='capture.png')
        program_nameS=Label(self.root,image=self.img00,font=('arial'),bg='black',fg='white',)
        program_nameS.place(x=0,y=0)
        menus=Menu(self.root)
        submenu0=Menu(menus)
        root.config(menu=menus)
        menus.add_cascade(label='File',menu=submenu0)
        submenu0.add_command(label='Exit',command=self.root.quit)
        self.loginframe=Frame(self.root,width=200, height=220,background="lightblue",)
        self.loginframe.config(bg=_from_rgb((204, 214, 229)))
        self.loginframe.place(x=500, y=200, width=340, height=200)
        self.p1=PhotoImage(file='img0.png')
        labels=Label(self.loginframe,image=self.p1)
        labels.place(x=0,y=0)
        login_label0=Label(self.loginframe,text='please enter your login credentials',font=('Garamond'),fg='black')
        login_label0.config(bg=_from_rgb((204, 214, 229)))
        login_label0.pack()
        login_label1=Label(self.loginframe,text='Username:',font=('Garamond'),bg=_from_rgb((204, 214, 229)))
        login_label1.place(x=0,y=45)
        login_label2=Label(self.loginframe,text='password:',font=('Garamond'),bg=_from_rgb((204, 214, 229)))
        login_label2.place(x=0,y=85)
        username=StringVar()
        password=StringVar()
        entry0a=Entry(self.loginframe,width='30', textvariable=username)
        entry0a.place(x=87,y=50)
        entry0b=Entry(self.loginframe,width='30',textvariable=password)
        entry0b.place(x=87,y=90,)
        buton0=Button(self.loginframe,text='Login',width=6,height=1,command=lambda:self.auth(str(username.get()),str(password.get())))
        buton0["border"] = "5"
        buton0.place(relx=.5, rely=.75, anchor="center")
        self.label3=Label(self.loginframe,text='Invalid credentials! please Try again',width=30,height=1,fg='black',bg='red')
        self.footer=Frame(self.root)
        self.footer.place(x=0,y=740,height=30,width=1300)
        self.img1t=PhotoImage(file='f.png')
        self.img2=PhotoImage(file='t.png')
        self.img3=PhotoImage(file='g.png')
        buttonN=Button(self.footer,image=self.img1t,command=lambda:openlink('https://www.facebook.com/Aschool.php'),relief=RIDGE)
        buttonN.place(relx=.973,rely=0)
        buttonN=Button(self.footer,image=self.img2,command=lambda:openlink('https://www.Twitter.com/Aschool'),relief=RIDGE)
        buttonN.place(relx=.946,rely=0)
        buttonN=Button(self.footer,image=self.img3,command=lambda:openlink('https://www.gmail.com/Aschool'),relief=RIDGE)
        buttonN.place(relx=.919,rely=0)
        self.img1=PhotoImage(file='h.png')
        buton0p=Button(self.root,image=self.img1,command=self.runa)
        buton0p["border"] = "4"
        buton0p.place(x=21,y=145, anchor="center")
        Thread(target=self.times).start()
        Thread(target=self.anim).start()
    def runa(self):
    	for widget in self.root.winfo_children():
    		widget.destroy()
    	TK=tk(self.root)
    def times(self):
	    while True:
	    	try:
	    		foot=Label(self.footer,text=time.strftime("%d/%m/%Y - %H:%M:%S"),width=23,height=1,anchor=W,font=('Agency FB','15'))
	    		foot.place(x=0,y=0)
	    		time.sleep(1)
    		except:
    			break
    def anim(self):
    	canvas = Canvas(self.footer,width = 1185, height = 20)
    	width=1190
    	canvas.place(x=0,y=0)
    	x=0
    	canvas.create_text(x, 10,text = "Allied High School®", tags = "text",font=('TimesNewRoman','15'))
    	dx=4
    	count=0
    	while True:
    		try:
	    		canvas.move('text',dx,0)
	    		canvas.after(30)
	    		canvas.update()
	    	except:
	    		break
    		if x<width:
    			x+=dx
    		else:
    			x=0
    			canvas.delete('text')
    			if count%2!=0:
    				canvas.create_text(x, 10,text = "Allied High School®",tags = "text",font=('TimesNewRoman','15'))
    				count=2
    			else:
    				canvas.create_text(x, 10,text = "አላይድ ሁለተኛ ደረጃ ትምህርት በት",tags = "text",font=('Times NewRoman','15'))
    				count+=1
    def auth(self,a,b):
    	db= openpyxl.load_workbook('data.xlsx')
    	sheet=db['Username and password']
    	row=sheet.max_row
    	for i in range(2,row+1):
	    	if a!='' and a==(sheet.cell(i,1).value) and str(b)==str(sheet.cell(i,2).value):#Adminstrators
	    		self.root=studentpage(self.root,a)
	    		self.loginframe.destroy()
	    		break
	    	elif a!='' and a==(sheet.cell(i,3).value) and b==(sheet.cell(i,4).value):#Adminstrators
	    		self.root=adminpage(self.root,a,((sheet.cell(i,5).value)))
	    		self.loginframe.destroy()
	    		break
	    	else:
		    	self.label3=Label(self.loginframe,width=30,height=1,fg='red',bg=_from_rgb((204, 214, 229)))
		    	self.label3.config(font=('Garamond',16),text='Invalid credentials!please Try again')
		    	self.label3.place(relx=0.0,y=170)

class adminpage(tk):
	def __init__(self,root,a,postion):
		super().__init__(root)
		self.loginframe.place_forget()
		self.postion=postion
		self.c=postion
		self.a=a
		self.root=root
		self.loginframe=Frame(self.root,width=200, height=220)
		self.p1=PhotoImage(file='img0.png')
		labels=Label(self.loginframe,image=self.p1)
		labels.place(x=0,y=0)
		self.program_name=Label(self.loginframe,text='Wellcome Mr.'+(self.a).capitalize(),font=('arial'),bg=_from_rgb((19,119,193)),fg='white',)
		self.program_name.pack(side=TOP,fill=X)
		self.loginframe.config(bg=_from_rgb((19,119,193)))
		self.loginframe.place(x=450, y=200, width=480, height=180)
		self.login_label0=Label(self.loginframe,text='please enter the username of the student you want to manage',font=('Garamond'),bg=_from_rgb((19,119,193)),fg='black',)
		self.login_label0.config(bg=_from_rgb((19,119,193)))
		self.login_label0.pack(side=TOP,fill=X)
		self.login_label2=Label(self.loginframe,text='Username:',font=('Garamond'),bg=_from_rgb((19,119,193)))
		self.login_label2.place(x=0,y=65)
		password=StringVar()
		entry0b=Entry(self.loginframe,width='30',textvariable=password)
		entry0b.place(x=92,y=68,)
		buton0=Button(self.loginframe,text='Go',width=6,height=1,command=lambda:self.stuselector(str(password.get())))
		buton0["border"] = "5"
		buton0.place(x=150,y=100)
		buton02=Button(self.loginframe,text='Click Here to register a new student',width=30,height=1,font=('Garamond',14),command=self.sturegister,bg=_from_rgb((19,119,193)))
		buton02["border"] = "1"
		buton02.place(x=0,y=145)
	def sturegister(self):
		for widget in self.loginframe.winfo_children():
			widget.destroy()
		self.img22=PhotoImage(file='b.png')
		buton0p=Button(self.root,image=self.img22,command=self.runa0)
		buton0p["border"] = "4"
		buton0p.place(x=21,y=176, anchor="center")
		self.p1=PhotoImage(file='img0.png')
		self.labels=Label(self.loginframe,image=self.p1)
		self.labels.place(x=0,y=0)
		self.login_label0=Label(self.loginframe,text='Please enter a Username for the student you are registering',font=('Garamond'),bg=_from_rgb((19,119,193)),fg='black',)
		self.login_label0.config(bg=_from_rgb((19,119,193)))
		self.login_label0.pack(side=TOP,fill=X)
		self.login_label2=Label(self.loginframe,text='Fullname:',font=('Garamond'),bg=_from_rgb((19,119,193)))
		self.login_label2.place(x=0,y=65)
		password=StringVar()
		entry0b=Entry(self.loginframe,width='30',textvariable=password)
		entry0b.place(x=92,y=68,)
		login_label22=Label(self.loginframe,text='Sex:',font=('Garamond'),bg=_from_rgb((19,119,193)))
		login_label22.place(x=0,y=95)
		password2=StringVar(value='Please fill as M or F')
		entry0b2=Entry(self.loginframe,width='30',textvariable=password2)
		entry0b2.place(x=92,y=98,)
		buton0=Button(self.loginframe,text='Register',width=6,height=1,command=lambda:self.confirmregister(str(password.get()),str(password2.get())))
		buton0["border"] = "5"
		buton0.place(relx=0.7,rely=.5)
	def confirmregister(self,b,c):
		self.b=b
		self.c=c
		db= openpyxl.load_workbook('data.xlsx')
		sheet=db['Username and password']
		row=sheet.max_row+1
		sheets=['Username and password','English','Mathematics','Historyy','Civics','Geography','ICT','HPE','Physics','Biology','Chemistry',]
		count=True
		for i in range(2,row+1):
			if self.b=='' or self.b == (sheet.cell(i,1).value) or self.c=='' or (self.c!='M' and self.c!='F'):
				self.label3=Label(self.loginframe,text='Student alredy exists or invalid input',width=30,height=1,fg='red')
				self.label3.config(font=('Garamond',16))
				self.label3.place(x=20,relx=0.0,rely=0.825)
				count=False
				break
		if count:
			for i in sheets:
				if i=='Username and password':
					pass
					sh=db[i]
					sh.cell(row,1).value=self.b
					x=random.randint(1000,10000)
					sh.cell(row,2).value=x
					tkinter.messagebox.showinfo("showinfo",('The new password for the user\t'+str(self.b)+'\tis\t'+str(x)))
					self.label3=Label(self.loginframe,text='Records updated',width=30,height=1,fg='Green')
					self.label3.config(font=('Garamond',16))
					self.label3.place(x=20,relx=0.0,rely=0.825)
				else:
					sh=db[i]
					sh.cell(row,2).value=self.b
					sh.cell(row,3).value=self.c
					sh.cell(row,1).value=row
					sh.cell(row,10).value=str('=SUM(D%s:I%s)'%(row,row))
		db.save('data.xlsx')
		time.sleep(1)
		self.sturegister()
	def stuselector(self,b):
		self.b=b
		db= openpyxl.load_workbook('data.xlsx')
		sheet=db['Username and password']
		row=sheet.max_row
		for i in range(2,row+1):
			if self.b!='' and self.b == (sheet.cell(i,1).value):
				self.root=adminmodify(self.root,self.b,self.postion)
				break
			else:
				self.label3=Label(self.loginframe,text='Student not found!',width=30,height=1,fg='red')
				self.label3.config(font=('Garamond',16))
				self.label3.place(x=20,relx=0.0,rely=0.825)
	def runa0(self):
		for widget in self.root.winfo_children():
			widget.destroy()
		TK=adminpage(self.root,self.a,self.c)
class adminmodify(adminpage):
	def __init__(self,root,a,c):
		super().__init__(root,a,c)
		self.loginframe.place_forget()
		self.c=c
		self.a=a
		self.root=root
		self.loginframe=Frame(self.root,width=200, height=220,background="lightblue",)
		self.loginframe.config(bg=_from_rgb((204, 214, 229)))
		self.loginframe.place(x=500, y=200, width=340, height=280)
		self.p1=PhotoImage(file='img0.png')
		labels=Label(self.loginframe,image=self.p1)
		labels.place(x=0,y=0)
		program_name=Label(self.loginframe,text=a,font=('arial'),bg='black',fg='white',)
		program_name.pack(side=TOP,fill=X)
		self.grade_modifier(a,c)
	def grade_modifier(self,a,c):
		self.img22=PhotoImage(file='b.png')
		buton0p=Button(self.root,image=self.img22,command=self.runa0)
		buton0p["border"] = "4"
		buton0p.place(x=21,y=176, anchor="center")
		db= openpyxl.load_workbook('data.xlsx',data_only=True)
		self.c=c
		self.a=a
		sheet=db[self.c]
		cell=sheet.cell(1, column = 2)
		row=sheet.max_row
		for i in range (2,row+1):
			if a in sheet.cell(i,2).value:
				c_row=i
				break
		#############  ####################  ##########################  ##########################  #######################  ##################
		login_label0=Label(self.loginframe,text=c,font=('Garamond'),bg='grey',fg='black',)
		login_label0.config(bg=_from_rgb((19,119,193)))
		login_label0.place(y=30,relx=.0)
		login_label2=Label(self.loginframe,text='Quiz1 10%:',font=('Garamond',12),bg=_from_rgb((19,119,193)))
		login_label2.place(x=10,y=74)
		Quiz1=StringVar(value=sheet.cell(c_row,4).value)
		Quiz1E=Entry(self.loginframe,width='5',textvariable=Quiz1,)
		Quiz1E.place(x=142,y=74,)
		####################################################################################################################
		login_label2=Label(self.loginframe,text='Assignment 1 10%:',font=('Garamond',12),bg=_from_rgb((19,119,193)))
		login_label2.place(x=10,y=94)
		Assignment1=StringVar(value=sheet.cell(c_row,5).value)
		Assignment1E=Entry(self.loginframe,width='5',textvariable=Assignment1,)
		Assignment1E.place(x=142,y=94,)
		####################################################################################################################
		login_label2=Label(self.loginframe,text='Mid 20%:',font=('Garamond',12),bg=_from_rgb((19,119,193)))
		login_label2.place(x=10,y=113)
		Mid=StringVar(value=sheet.cell(c_row,6).value)
		MidE=Entry(self.loginframe,width='5',textvariable=Mid,)
		MidE.place(x=142,y=113,)
		####################################################################################################################
		login_label2=Label(self.loginframe,text='Quiz2 10%:',font=('Garamond',12),bg=_from_rgb((19,119,193)))
		login_label2.place(x=10,y=131)
		Quiz2=StringVar(value=sheet.cell(c_row,7).value)
		Quiz2E=Entry(self.loginframe,width='5',textvariable=Quiz2,)
		Quiz2E.place(x=142,y=131,)
		####################################################################################################################
		login_label2=Label(self.loginframe,text='Assignment2 10%:',font=('Garamond',12),bg=_from_rgb((19,119,193)))
		login_label2.place(x=10,y=150)
		Assignment2=StringVar(value=sheet.cell(c_row,8).value)
		Assignment2E=Entry(self.loginframe,width='5',textvariable=Assignment2,)
		Assignment2E.place(x=142,y=150,)
		####################################################################################################################
		login_label2=Label(self.loginframe,text='Final 40%',font=('Garamond',12),bg=_from_rgb((19,119,193)))
		login_label2.place(x=10,y=169)
		Final2=StringVar(value=sheet.cell(c_row,9).value)
		Final2E=Entry(self.loginframe,width='5',textvariable=Final2,)
		Final2E.place(x=142,y=169,)
		####################################################################################################################
		login_label2=Label(self.loginframe,text='Total',font=('Garamond',12),bg=_from_rgb((19,119,193)))
		login_label2.place(x=10,y=188)
		try:
			Total2=StringVar(value=self.total(eval(Quiz1E.get()),eval(Assignment1E.get()),eval(MidE.get()),eval(Quiz2E.get()),eval(Assignment2E.get()),eval(Final2E.get())))
		except:
			Total2=StringVar(value=self.total(0,0,0,0,0,0))
		Total2E=Entry(self.loginframe,width='5',textvariable=Total2,state='disabled')
		Total2E.place(x=142,y=188,)
		login_label0=Label(self.loginframe,text=a,font=('Garamond'),bg='grey',fg='black',anchor=W)
		try:
			buton0=Button(self.loginframe,text='update',width=6,height=1,command=lambda:self.dataupdate(str(Quiz1E.get()),str(Assignment1E.get()),str(MidE.get()),str(Quiz2E.get()),str(Assignment2E.get()),str(Final2E.get()),self.c,self.a))
			buton0["border"] = "5"
			buton0.place(x=71,y=214)
		except:
			buton0=Button(self.loginframe,text='update',width=6,height=1,command=lambda:self.dataupdate(0,0,0,0,0,0,self.c,self.a))
			buton0["border"] = "5"
			buton0.place(x=71,y=214)
	def total(self,a,b,c,d,e,f):
		return a+b+c+d+e+f
	def dataupdate(self,a,b,c,d,e,f,sheet,dudesname):
		db= openpyxl.load_workbook('data.xlsx')
		self.c=sheet
		sheet=db[self.c]
		row=sheet.max_row
		for i in range (2,row+1):
			if dudesname in sheet.cell(i,2).value:
				c_row=i
				break
		sheet.cell(row=c_row,column=4).value =int(a)
		sheet.cell(row=c_row,column=5).value =int(b)
		sheet.cell(row=c_row,column=6).value =int(c)
		sheet.cell(row=c_row,column=7).value =int(d)
		sheet.cell(row=c_row,column=8).value =int(e)
		sheet.cell(row=c_row,column=9).value =int(f)
		time.sleep(2)
		self.label3=Label(self.loginframe,text='Records have been updated',width=30,height=1,fg='Green',bg=_from_rgb((204, 214, 229)))
		self.label3.config(font=('Garamond',16))
		self.label3.place(relx=0.0,rely=0.9)
		db.save('data.xlsx')
		self.grade_modifier(self.a,self.c)
	def runa0(self):
		#self.loginframe.place_forget()
		for widget in self.root.winfo_children():
			widget.destroy()
		TK=adminpage(self.root,self.a,self.c)
class studentpage(tk):
	def __init__(self,root,a):
		super().__init__(root)
		self.loginframe.place_forget()
		self.root=root
		self.a=a
		self.loginframe=Frame(self.root,width=400, height=420,)
		#self.loginframe.config(bg=_from_rgb((255, 255, 255)))
		self.loginframe.place(relx=0.25, y=130, width=700, height=325)
		self.p1=PhotoImage(file='img0.png')
		labels=Label(self.loginframe,image=self.p1)
		labels.place(x=0,y=0)
		program_name=Label(self.loginframe,text='Wellcome Mr.'+(self.a).capitalize(),font=('arial'),fg='black',bg=_from_rgb((19,119,193)))
		program_name2=Label(self.loginframe,text='Please Select a Subject',font=('arial'),fg='black',bg=_from_rgb((19,119,193)))
		program_name.pack(side=TOP,fill=X)
		program_name2.pack(side=TOP,fill=X)
		self.selector=Frame(self.loginframe,width=400, height=325)
		self.selector.config(bg=_from_rgb((19,119,193)))
		self.selector.place(relx=.0, rely=0, width=40, height=180)
		self.label_frame=Frame(self.loginframe)
		buton0=Button(self.loginframe,text='English',width=9,height=1,command=lambda:self.grade_reader('English',self.a),bg=_from_rgb((19,119,193)))
		buton0["border"] = "5"
		buton0.place(relx=0.0,y=55)
		buton0=Button(self.loginframe,text='Mathematics',width=9,height=1,command=lambda:self.grade_reader('Mathematics',self.a),bg=_from_rgb((19,119,193)))
		buton0["border"] = "5"
		buton0.place(relx=0.0,y=55+30)
		buton0=Button(self.loginframe,text='Physics',width=9,height=1,command=lambda:self.grade_reader('Physics',self.a),bg=_from_rgb((19,119,193)))
		buton0["border"] = "5"
		buton0.place(relx=0.0,y=55+60)
		buton0=Button(self.loginframe,text='Biology',width=9,height=1,command=lambda:self.grade_reader('Biology',self.a),bg=_from_rgb((19,119,193)))
		buton0["border"] = "5"
		buton0.place(relx=0.0,y=55+90)
		buton0=Button(self.loginframe,text='Chemistry',width=9,height=1,command=lambda:self.grade_reader('Chemistry',self.a),bg=_from_rgb((19,119,193)))
		buton0["border"] = "5"
		buton0.place(relx=0.0,y=55+120)
		buton0=Button(self.loginframe,text='Geography',width=9,height=1,command=lambda:self.grade_reader('Geography',self.a),bg=_from_rgb((19,119,193)))
		buton0["border"] = "5"
		buton0.place(relx=0.0,y=55+150)
		buton0=Button(self.loginframe,text='History',width=9,height=1,command=lambda:self.grade_reader('Historyy',self.a),bg=_from_rgb((19,119,193)))
		buton0["border"] = "5"
		buton0.place(relx=0.0,y=55+180)
		buton0=Button(self.loginframe,text='Civics',width=9,height=1,command=lambda:self.grade_reader('Civics',self.a),bg=_from_rgb((19,119,193)))
		buton0["border"] = "5"
		buton0.place(relx=0.0,y=55+210)
		buton0=Button(self.loginframe,text='ICT',width=9,height=1,command=lambda:self.grade_reader('ICT',self.a),bg=_from_rgb((19,119,193)))
		buton0["border"] = "5"
		buton0.place(relx=0.0,y=55+240)
		buton0=Button(self.loginframe,text='HPE',width=9,height=1,command=lambda:self.grade_reader('HPE',self.a),bg=_from_rgb((19,119,193)))
		buton0["border"] = "5"
		buton0.place(relx=0.0,y=55+270)
	def grade_reader(self,c,a):
		self.label_frame.destroy()
		db= openpyxl.load_workbook('data.xlsx',data_only=True)
		self.c=c
		self.a=a
		sheet=db[self.c]
		row=sheet.max_row
		for i in range (2,row+1):
			if a in sheet.cell(i,2).value:
				c_row=i
				break
		#############  ####################  ##########################  ##########################  #######################  ##################
		self.label_frame=Frame(self.loginframe)
		self.label_frame.place(y=30,relx=.0)
		login_label0=Label(self.label_frame,text=c,font=('Garamond'),fg='black',bg=_from_rgb((0,100,177)))
		login_label0.pack()
		login_label2=Label(self.loginframe,text='Quiz1 10%:',font=('Garamond',12),bg=_from_rgb((0,100,177)))
		login_label2.place(x=200+20,y=40+74)
		Quiz1=StringVar(value=sheet.cell(c_row,4).value)
		Quiz1E=Entry(self.loginframe,width='5',state='disabled',textvariable=Quiz1,)
		Quiz1E.place(x=200+242,y=40+74,)
		####################################################################################################################
		login_label2=Label(self.loginframe,text='Assignment 1 10%:',font=('Garamond',12),bg=_from_rgb((0,100,177)))
		login_label2.place(x=200+20,y=40+94)
		Assignment1=StringVar(value=sheet.cell(c_row,5).value)
		Assignment1E=Entry(self.loginframe,width='5',state='disabled',textvariable=Assignment1,)
		Assignment1E.place(x=200+242,y=40+94,)
		####################################################################################################################
		login_label2=Label(self.loginframe,text='Mid 20%:',font=('Garamond',12),bg=_from_rgb((0,100,177)))
		login_label2.place(x=200+20,y=40+113)
		Mid=StringVar(value=sheet.cell(c_row,6).value)
		MidE=Entry(self.loginframe,width='5',state='disabled',textvariable=Mid,)
		MidE.place(x=200+242,y=40+113,)
		####################################################################################################################
		login_label2=Label(self.loginframe,text='Quiz2 10%:',font=('Garamond',12),bg=_from_rgb((0,100,177)))
		login_label2.place(x=200+20,y=40+131)
		Quiz2=StringVar(value=sheet.cell(c_row,7).value)
		Quiz2E=Entry(self.loginframe,width='5',state='disabled',textvariable=Quiz2,)
		Quiz2E.place(x=200+242,y=40+131,)
		####################################################################################################################
		login_label2=Label(self.loginframe,text='Assignment2 10%:',font=('Garamond',12),bg=_from_rgb((0,100,177)))
		login_label2.place(x=200+20,y=40+150)
		Assignment2=StringVar(value=sheet.cell(c_row,8).value)
		Assignment2E=Entry(self.loginframe,width='5',state='disabled',textvariable=Assignment2,)
		Assignment2E.place(x=200+242,y=40+150,)
		####################################################################################################################
		login_label2=Label(self.loginframe,text='Final 40%',font=('Garamond',12),bg=_from_rgb((0,100,177)))
		login_label2.place(x=200+20,y=40+169)
		Final2=StringVar(value=sheet.cell(c_row,9).value)
		Final2E=Entry(self.loginframe,width='5',state='disabled',textvariable=Final2,)
		Final2E.place(x=200+242,y=40+169,)
		####################################################################################################################
		login_label2=Label(self.loginframe,text='Total',font=('Garamond',12),bg=_from_rgb((0,100,177)))
		login_label2.place(x=200+20,y=40+188)
		try:
			Total2=StringVar(value=self.total(eval(Quiz1E.get()),eval(Assignment1E.get()),eval(MidE.get()),eval(Quiz2E.get()),eval(Assignment2E.get()),eval(Final2E.get())))
		except:
			Total2=StringVar(value=self.total(0,0,0,0,0,0))
		Total2E=Entry(self.loginframe,width='5',state='disabled',textvariable=Total2)
		Total2E.place(x=200+242,y=40+169+20)
		login_label0=Label(self.loginframe,text=a,font=('Garamond'),fg='black',bg=_from_rgb((0,100,177)))	
	def total(self,a,b,c,d,e,f):
		return a+b+c+d+e+f
roots=Tk()
root=tk(roots)
roots.mainloop()
